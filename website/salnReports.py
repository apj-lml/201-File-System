from flask import Blueprint, request, redirect, url_for, session, jsonify, current_app, send_file
from flask_login import current_user, login_required
from flask_principal import Permission, RoleNeed
from .models import Real_Property, Personal_Property, Liability, Business_Interest, Relatives_In_Government,\
     User, UserSchema, PersonalPropertySchema, Real_Property, Personal_Property, Saln_Summary, Liability, RealPropertySchema, LiabilitySchema, BusinessInterestSchema, RelativeInGovernmentSchema
from . import db
import datetime
import time
from .myhelper import allowed_file, my_random_string
from werkzeug.utils import secure_filename
from io import BytesIO
from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from dateutil.relativedelta import relativedelta
from sqlalchemy import func, or_
from sqlalchemy.sql import label 
import json
import os, os.path
from openpyxl import load_workbook
import xlsxwriter
# from openpyxl.styles import Border, Side
# from editpyxl import Workbook
# from editpyxl.styles import Border
# from openpyxl_templates import TemplatedWorkbook

salnReports = Blueprint('salnReports', __name__)
# ALLOWED_EXTENSIONS = {'pdf'}

admin_permission = Permission(RoleNeed('admin'))

@salnReports.errorhandler(403)
def page_not_found(e):
    session['redirected_from'] = request.url
    return redirect(url_for('auth.login'))


# ---------------------------------------------------------------------------- #
#                                PRINT salnReports                                    #
# ---------------------------------------------------------------------------- #
@salnReports.route('print/<emp_id>/<filing_date>/<filing_type>', methods=['POST', 'GET'])
@login_required
def print_salnReports(emp_id, filing_date, filing_type):
    # formdata  = json.loads(request.data)
    template = os.path.join(current_app.root_path, 'static/templates', 'salnReports_Form.docx')   

    document = from_template(template, emp_id, filing_date, filing_type)
    document.seek(0)
    
    return send_file(
        document, mimetype='application/vnd.openxmlformats-'
        'officedocument.wordprocessingml.document', as_attachment=True,
        attachment_filename='salnReports.docx')

 # ---------------------------------------------------------------------------- #

def from_template(template, emp_id, filing_date, filing_type):
    target_file = BytesIO()
    
    template = DocxTemplate(template)
    context = get_context(emp_id, filing_date, filing_type)  # gets the context used to render the document
    
    target_file = BytesIO()
    template.render(context, autoescape=True)
    template.save(target_file)
    return target_file

def get_context(id, filing_date, filing_type):
    user = db.session.query(User).get(id)
    
    user_profile = user

    # user_profile.employee_id_date_issued = user_profile.employee_id_date_issued.strftime("%B %d, %Y")

    user_profile.user_real_property = sorted(user_profile.user_real_property, key=lambda x: x.rp_acquisition_year, reverse=True)
    user_profile.user_personal_property = sorted(user_profile.user_personal_property, key=lambda x: x.pp_year_acquired, reverse=True)

    middle_name = user_profile.middle_name[:1] + "." if user_profile.middle_name and user_profile.middle_name != "N/A" else ""
    name_extn = user_profile.name_extn if user_profile.name_extn and user_profile.name_extn != "N/A" else ""

    bi_list = []
    for bi in user_profile.user_business_interest:
        bi_acquistion_date_object = datetime.datetime.strptime(bi.business_acquisition, "%Y-%m-%d").date()
        bi.business_acquisition = bi_acquistion_date_object.strftime("%B %d, %Y")
        bi_list.append(bi)

    user_profile.user_business_interest = bi_list

    rg_list = []
    for rg in user_profile.user_relative_in_government:
        # rg_acquistion_date_object = datetime.datetime.strptime(rg.business_acquisition, "%Y-%m-%d").date()
        # rg.business_acquisition = rg_acquistion_date_object.strftime("%B %d, %Y")
        rg_list.append(rg)

    user_profile.user_relative_in_government = rg_list

    user_profile_dict = UserSchema().dump(user_profile)

    user_profile_dict['full_name'] = user_profile.first_name + " " + middle_name + " " + user_profile.name + " " + name_extn
    address = user_profile.p_house_block_lot + ", " + user_profile.p_street + ", " + user_profile.p_subdivision_village + ", " + user_profile.p_barangay + ", " + user_profile.p_city_municipality + ", "+ user_profile.p_province + ", "+ user_profile.p_zip_code
    final_address = address.replace('N/A,', '')
    user_profile_dict['address'] = " ".join(final_address.split())

    user_profile_dict_date_object = datetime.datetime.strptime(user_profile_dict['employee_id_date_issued'], "%Y-%m-%d").date()
    user_profile_dict['employee_id_date_issued'] = user_profile_dict_date_object.strftime("%B %d, %Y")

    # user_profile_dict['spouse_name'] = user_profile.familyBg.filter_by(fb_relationship='SPOUSE').first()

    children_list = []

    for child in user_profile.familyBg:
        if child.fb_relationship == 'CHILD':
            filing_date_object = datetime.datetime.strptime(filing_date, "%Y-%m-%d").date()
            dob_object = datetime.datetime.strptime(child.fb_date_of_birth, "%Y-%m-%d").date()
            age_based_on_input = relativedelta(filing_date_object, dob_object)

            if age_based_on_input.years < 18:
                if age_based_on_input.years <= 0:
                    child.childAge = str(age_based_on_input.months) + " MONTH/S OLD"
                else:
                    child.childAge = str(age_based_on_input.years)


                child.formattedBirthDate = dob_object.strftime("%B %d, %Y")
                children_list.append(child)

    # Sort children_list in descending order based on child's age
    # children_list = sorted(children_list, key=lambda x: x.childAge, reverse=True)

    spouse = None
    for relative in user_profile.familyBg:
        if relative.fb_relationship == 'SPOUSE':
            spouse = relative
            break

    user_profile_dict['spouse_name'] = spouse.fb_name if spouse else "N/A"
    user_profile_dict['spouse_first_name'] = spouse.fb_first_name if spouse else "N/A"
    user_profile_dict['spouse_middle_name'] = spouse.fb_middle_name if spouse else "N/A"
    user_profile_dict['spouse_middle_initial'] = spouse.fb_middle_name[0] + "." if spouse else "N/A"
    user_profile_dict['spouse_full_name'] = spouse.fb_first_name + " " + spouse.fb_middle_name[0] + ". " + spouse.fb_name if spouse else "N/A"
    # user_profile_dict['spouse_name'] = user_profile.familyBg.filter_by(fb_relationship='SPOUSE').first().name if user_profile.familyBg else None

    user_profile_dict['spouse_position'] = spouse.fb_occupation if spouse else "N/A"
    user_profile_dict['spouse_agency'] = spouse.fb_employer_business_name if spouse else "N/A"
    user_profile_dict['spouse_office_address'] = spouse.fb_business_address if spouse else "N/A"

    user_profile_dict['spouse_govt_issued_id'] = spouse.fb_id if spouse and spouse.fb_id != "" else "N/A"
    user_profile_dict['spouse_govt_issued_id_no'] = spouse.fb_id_no if spouse and spouse.fb_id_no != "" else "N/A"

    if spouse is not None:
        if spouse.fb_deceased == "unchecked":
            user_profile_dict['spouse_deceased'] = "unchecked"
        elif spouse.fb_deceased == "checked":
            user_profile_dict['spouse_deceased'] = "checked"

        if spouse.fb_abroad == "unchecked":
            user_profile_dict['spouse_abroad'] = "unchecked"
        elif spouse.fb_abroad == "checked":
            user_profile_dict['spouse_abroad'] = "checked"
    else:
        user_profile_dict['spouse_deceased'] = "unchecked"
        user_profile_dict['spouse_abroad'] = "unchecked"


    spouse_govt_issued_id_date_issued_date_obj = datetime.datetime.strptime(spouse.fb_date_issued, "%Y-%m-%d").date() if spouse and spouse.fb_date_issued != "" and spouse.fb_date_issued is not None else "N/A"

    user_profile_dict['spouse_govt_issued_id_date_issued'] = spouse_govt_issued_id_date_issued_date_obj.strftime("%B %d, %Y") if spouse and spouse.fb_date_issued != "" and spouse.fb_date_issued is not None else "N/A"

    user_profile_dict["checkmark"] = "✓"

    date_object = datetime.datetime.strptime(filing_date, "%Y-%m-%d")

    # Format the datetime object as "Month Day, Year"
    formatted_date = date_object.strftime("%B %d, %Y")

    user_profile_dict["filing_date"] = formatted_date
    user_profile_dict["filing_type"] = filing_type
    user_profile_dict["children_list"] = sorted(children_list, key=lambda x: x.fb_date_of_birth, reverse=False)

    user_profile_dict['total_rp_acquisition_cost_p1'] = formatNumber(getRpAcquisitionCostSubTotal(user, 0, 3) + getRpAcquisitionCostSubTotal(user, 3, 7)) if getRpAcquisitionCostSubTotal(user, 0, 3) != 0 else formatNumber(0.00)
    user_profile_dict['total_rp_acquisition_cost_p2'] = formatNumber(getRpAcquisitionCostSubTotal(user, 3, 7)) if getRpAcquisitionCostSubTotal(user, 3, 7) != 0 else formatNumber(0.00)

    user_profile_dict['total_pp_acquisition_cost_p1'] = formatNumber(getPpAcquisitionCostSubTotal(user, 0, 6) + getPpAcquisitionCostSubTotal(user, 6, 12)) if getPpAcquisitionCostSubTotal(user, 0, 6) != 0 else formatNumber(0.00)
    user_profile_dict['total_pp_acquisition_cost_p2'] = formatNumber(getPpAcquisitionCostSubTotal(user, 6, 12)) if getPpAcquisitionCostSubTotal(user, 6, 12) != 0 else formatNumber(0.00)

    user_profile_dict['total_liability_outstanding_balance_p1'] = formatNumber(getLiabilityOutstandingBalance(user, 0, 3) + getLiabilityOutstandingBalance(user, 3, 8)) if getLiabilityOutstandingBalance(user, 0, 3) != 0 else formatNumber(0.00)
    user_profile_dict['total_liability_outstanding_balance_p2'] = formatNumber(getLiabilityOutstandingBalance(user, 3, 8)) if getLiabilityOutstandingBalance(user, 3, 8) != 0 else formatNumber(0.00)
    
    user_profile_dict['total_assets_p1'] = formatNumber(getTotalAssets(getRpAcquisitionCostSubTotal(user, 0, 3), getPpAcquisitionCostSubTotal(user, 0, 6)) + getTotalAssets(getRpAcquisitionCostSubTotal(user, 3, 7), getPpAcquisitionCostSubTotal(user, 6, 12)))
    user_profile_dict['total_assets_p2'] = formatNumber(getTotalAssets(getRpAcquisitionCostSubTotal(user, 3, 7), getPpAcquisitionCostSubTotal(user, 6, 12)))

    user_profile_dict['networth'] = formatNumber(getNetworth(user_profile_dict['total_assets_p1'], user_profile_dict['total_liability_outstanding_balance_p1'], '0.00', '0.00'))

    user_profile_dict['signatory'] = user.assignatory[0].assignatory
    user_profile_dict['signatory_position_title'] = user.assignatory[0].position_title

    # if getRpAcquisitionCostSubTotal(user, 3, 7) != None or getPpAcquisitionCostSubTotal(user, 6, 12) != None or getLiabilityOutstandingBalance(user, 3, 8) != None or len(user_profile.user_business_interest) > 2:
    if getRpAcquisitionCostSubTotal(user, 3, 8) != 0.00 or getPpAcquisitionCostSubTotal(user, 6, 13) != 0.00 or getLiabilityOutstandingBalance(user, 3, 9) != 0.00:
        user_profile_dict['addtl_page'] = True
    else:
        user_profile_dict['addtl_page'] = False

    if len(user_profile.user_business_interest) > 2 or len(user_profile.user_relative_in_government) > 5:
        user_profile_dict['addtl_page_3'] = True
    else:
        user_profile_dict['addtl_page_3'] = False

    return user_profile_dict

def getNetworth(total_assets_p1 = 0.00, total_liability_p1 = 0.00, total_assets_p2 = 0.00, total_liability_p2 = 0.00):
    # if total_assets_p1:
    #     floatAssetsP1 = float(total_assets_p1.replace(',', ''))
    # else:
    #     floatAssetsP1 = 0.00

    # if total_assets_p2:
    #     floatAssetsP2 = float(total_assets_p2.replace(',', ''))
    # else:
    #     floatAssetsP2 = 0.00

    # if total_liability_p1:
    #     floatLiabilityP1 = float(total_liability_p1.replace(',', ''))
    # else:
    #     floatLiabilityP1 = 0.00

    # if total_liability_p2:
    #     floatLiabilityP2 = float(total_liability_p2.replace(',', ''))
    # else:
    #     floatLiabilityP2 = 0.00

    floatNetworth = float(total_assets_p1.replace(',', '')) + float(total_assets_p2.replace(',', '')) - float(total_liability_p1.replace(',', '')) + float(total_liability_p2.replace(',', ''))
    # floatNetworth = (floatAssetsP1 + floatAssetsP2) - (floatLiabilityP1 + floatLiabilityP2)
    formatted_networth = "{:,.2f}".format(floatNetworth)
    return floatNetworth

def getTotalAssets(subtotal1 = 0.00, subtotal2 = 0.00):
    # if subtotal1:
    #     floatSubtotal1 = float(subtotal1.replace(',', ''))
    # else:
    #     floatSubtotal1 = 0.00

    # if subtotal2:
    #     floatSubtotal2 = float(subtotal2.replace(',', ''))
    # else:
    #     floatSubtotal2 = 0.00

    # total = floatSubtotal1 + floatSubtotal2
    total = subtotal1 + subtotal2
    formatted_total = "{:,.2f}".format(total)

    return total
    # return formatted_total

def getLiabilityOutstandingBalance(user, d_start, d_end):
    # Extract the acquisition_cost values from the first four real properties (if available)
    outstanding_balances = []
    liabilities = user.user_liability
    if liabilities:
        for liability in liabilities[d_start:d_end]:
            liability_str = liability.liability_outstanding_balance
            if liability_str:
                # Remove commas and convert to a float
                outstanding_balance = float(liability_str.replace(',', ''))
                outstanding_balances.append(outstanding_balance)
            else:
                formatted_total_outstanding_balance = None


        # Calculate the sum of the first four acquisition_cost values
        total_outstanding_balance = sum(outstanding_balances) if sum(outstanding_balances) > 0 else 0.00
        formatted_total_outstanding_balance = "{:,.2f}".format(total_outstanding_balance) if sum(outstanding_balances) > 0 else None
    else:
        formatted_total_outstanding_balance = None
        total_outstanding_balance = 0.00
    
    return total_outstanding_balance
    # return formatted_total_outstanding_balance

def formatNumber(x):
        return "{:,.2f}".format(x)
     
def getRpAcquisitionCostSubTotal(user, d_start, d_end):
    # Extract the acquisition_cost values from the first four real properties (if available)
    acquisition_costs = []
    # real_properties = user.user_real_property
    real_properties = sorted(user.user_real_property, key=lambda x: x.rp_acquisition_year, reverse=True)

    if real_properties:
        for real_property in real_properties[d_start:d_end]:
            acquisition_cost_str = real_property.rp_acquisition_cost
            if acquisition_cost_str:
                # Remove commas and convert to a float
                acquisition_cost = float(acquisition_cost_str.replace(',', ''))
                acquisition_costs.append(acquisition_cost)
            else:
                formatted_total_acquisition_cost = 0

        # Calculate the sum of the first four acquisition_cost values
        total_acquisition_cost = sum(acquisition_costs) if sum(acquisition_costs) > 0 else 0.00
        formatted_total_acquisition_cost = "{:,.2f}".format(total_acquisition_cost) if sum(acquisition_costs) > 0 else None
    else:
        formatted_total_acquisition_cost = 0
        total_acquisition_cost = 0
    
    return total_acquisition_cost
    # return formatted_total_acquisition_cost


def getPpAcquisitionCostSubTotal(user, d_start, d_end):
    # Extract the acquisition_cost values from the first four real properties (if available)
    acquisition_costs = []
    personal_properties = user.user_personal_property
    # personal_properties = sorted(user.user_personal_property, key=lambda x: x.pp_year_acquired, reverse=True)

    if personal_properties:
        for personal_property in personal_properties[d_start:d_end]:
            acquisition_cost_str = personal_property.pp_acquisition_cost
            if acquisition_cost_str:
                # Remove commas and convert to a float
                acquisition_cost = float(acquisition_cost_str.replace(',', ''))
                acquisition_costs.append(acquisition_cost)
            else:
                formatted_total_acquisition_cost = None

        # Calculate the sum of the first four acquisition_cost values
        total_acquisition_cost = sum(acquisition_costs) if sum(acquisition_costs) > 0 else 0.00
        formatted_total_acquisition_cost = "{:,.2f}".format(total_acquisition_cost) if sum(acquisition_costs) > 0 else None
    else:
        formatted_total_acquisition_cost = None
        total_acquisition_cost = 0.00
    
    # return formatted_total_acquisition_cost
    return total_acquisition_cost

# ---------------------------------------------------------------------------- #

# ---------------------------------------------------------------------------- #
#                                     get data                                 #
# ---------------------------------------------------------------------------- #
@salnReports.route('get-data', methods=['POST', 'GET'])
@login_required
def get_data():
    user_profiles = User.query.all()


    return jsonify(user_list)
# ---------------------------------------------------------------------------- #


# ---------------------------------------------------------------------------- #
#                                     get data                                 #
# ---------------------------------------------------------------------------- #
@salnReports.route('generate-summary-list', methods=['POST', 'GET'])
@login_required
def generate_summary_list():
    users_with_saln = User.query.join(Saln_Summary).filter(User.saln_summary != None).filter(User.status_remarks == "ACTIVE").filter(User.employment_status != "JOB ORDER").filter(User.employment_status != "CONTRACT OF SERVICE").all()
    jo_users_with_saln = User.query.join(Saln_Summary).filter(
                                                            User.saln_summary != None,
                                                            User.status_remarks == "ACTIVE",
                                                            or_(User.employment_status == "JOB ORDER", User.employment_status == "CONTRACT OF SERVICE")
                                                            ).all()

    template = os.path.join(current_app.root_path, 'static/templates', 'summary_list_of_filers.xlsx')  
    new_save_path = os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers.xlsx')  

    # Create a BytesIO buffer to save the workbook
    buffer = BytesIO()

    # workbook = xlsxwriter.Workbook(os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers_test.xlsx'))
    workbook = xlsxwriter.Workbook(buffer)
    # default cell format to size 10 

    ws = workbook.add_worksheet("PERM-CAS-COT")
    ws.set_paper(9)
    ws.set_landscape()
    ws.center_horizontally()
    ws.set_margins(left=0.13, right=0.13, top=1.21, bottom=1.21)

    ws.repeat_rows(0, 4)


    ws.set_header('&L&G', {'image_left': os.path.join(current_app.root_path, 'static/saln', 'landscape-header-cropped.png')}, margin=0)
    ws.set_footer('&R&G', {'image_right': os.path.join(current_app.root_path, 'static/saln', 'landscape-footer-cropped.png')}, margin=0)

    cell_format = workbook.add_format({'bold': True, 'font_name':'Cambria'})
    cell_format_cambria = workbook.add_format({'font_name':'Cambria'})
    cell_format_cambria_total = workbook.add_format({'font_name':'Cambria', 'valign':'left', 'font_size':9})
    cell_format_cambria_contact = workbook.add_format({'font_name':'Cambria', 'valign':'left', 'font_size':9})
    cell_format_cambria_italic = workbook.add_format({'font_name':'Cambria', 'italic' : True,})
    cell_format_cambria_bold= workbook.add_format({'font_name':'Cambria', 'bold' : True,})
    cell_format_cambria_italic_signatory = workbook.add_format({'font_name':'Cambria', 'italic' : True, 'align': 'center', 'valign': 'left'})
    cell_format_cambria_bold_signatory= workbook.add_format({'font_name':'Cambria', 'bold' : True, 'align': 'center', 'valign': 'left'})
    cell_format_center = workbook.add_format({'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter'})
    cell_header_format = workbook.add_format({'bold': True, 'font_name':'Cambria', 'align': 'center', 'font_size': 14})
    cell_subheader = workbook.add_format({
                                        'font_name':'Cambria',
                                        'align':    'center',
                                        'valign':   'top',})
    table_header_format = workbook.add_format({'bold': True, 'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter', 'border' : 1 })
    table_subheader_format = workbook.add_format({'bold': True, 'italic' : True, 'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter', 'border' : 1})
    table_body = workbook.add_format({'valign': 'left','font_name':'Cambria', 'border' : 1, 'font_size': 9})
    table_body_a = workbook.add_format({'valign': 'center','font_name':'Cambria', 'border' : 1, 'font_size': 9})

    employee_status = workbook.add_format({'bold': True,'font_name':'Cambria', 'border' : 1, 'font_size': 9})

    ws.merge_range("A1:I1", "SUMMARY LIST OF FILERS", cell_header_format)
    ws.merge_range("A2:I2", "Statement of Assets, Liabilities and Net Worth", cell_subheader)
    ws.merge_range("A3:I3", "Calendar Year 2023", cell_subheader)

    ws.merge_range("A4:A5", "NO.", table_header_format)
    ws.merge_range("B4:E4", "NAME OF EMPLOYEE", table_header_format)
    ws.merge_range("B5:C5", "Last Name", table_subheader_format)
    ws.write("D5", "First Name", table_subheader_format)
    ws.write("E5", "Middle Name", table_subheader_format)
    ws.merge_range("F4:F5", "TIN", table_header_format)
    ws.merge_range("G4:H5", "POSITION TITLE", table_header_format)
    # ws.merge_range("G4:G5", "NET WORTH", table_header_format)
    ws.merge_range("I4:I5", "NET WORTH", table_header_format)

    ws.merge_range("A6:I6", "PERMANENT", employee_status)

    ws.set_row(0, 15)
    ws.set_row(1, 14.25)
    ws.set_row(2, 22.5)
    ws.set_row(3, 15, cell_format)

    ws.set_column('A:A', 4.29, cell_format_center)
    ws.set_column('B:B', 14.43, cell_format_cambria)
    ws.set_column('C:C', 0.83, cell_format_cambria)
    ws.set_column('D:D', 16.57, cell_format_cambria)
    ws.set_column('E:E', 15.29, cell_format_center)
    ws.set_column('F:F', 14.14, cell_format_cambria)
    ws.set_column('G:G', 14.86, cell_format_cambria)
    ws.set_column('H:H', 40.43, cell_format_cambria)
    ws.set_column('I:I', 19.29, cell_format_cambria)

    row = 7

    users_with_saln_and_permanent = [user for user in users_with_saln if user.employment_status == "PERMANENT"]
    for index, user_with_saln in enumerate(users_with_saln_and_permanent, start=0):
        ws.write(row-1, 0, index+1, table_body_a)
        ws.merge_range("B"+str(row) + ":C"+str(row), user_with_saln.last_name, table_body)
        ws.write(row-1, 3, f"{user_with_saln.first_name}, {user_with_saln.name_extn}" if user_with_saln.name_extn != 'N/A' else user_with_saln.first_name, table_body)
        ws.write(row-1, 4, user_with_saln.middle_name, table_body)
        ws.write(row-1, 5, user_with_saln.tin, table_body_a)
        ws.merge_range("G"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
        ws.write(row-1, 8, user_with_saln.saln_summary.networth, table_body_a)
        
        row = row + 1

    users_with_saln_and_coterm = [user for user in users_with_saln if user.employment_status == "COTERMINOUS"]
    
    if users_with_saln_and_coterm:
        ws.merge_range("A" + str(row) +":I" + str(row), "COTERMINOUS", employee_status)
        row = row + 1
        
        for index, user_with_saln in enumerate(users_with_saln_and_coterm, start=0):
            ws.write(row-1, 0, index+1, table_body_a)
            ws.merge_range("B"+str(row) + ":C"+str(row), user_with_saln.last_name, table_body)
            ws.write(row-1, 3, f"{user_with_saln.first_name}, {user_with_saln.name_extn}" if user_with_saln.name_extn != 'N/A' else user_with_saln.first_name, table_body)
            ws.write(row-1, 4, user_with_saln.middle_name, table_body)
            ws.write(row-1, 5, user_with_saln.tin, table_body_a)
            ws.merge_range("G"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
            ws.write(row-1, 8, user_with_saln.saln_summary.networth, table_body_a)
            
            row = row + 1

    users_with_saln_and_casual = [user for user in users_with_saln if user.employment_status == "CASUAL"]
    
    if users_with_saln_and_casual:
        ws.merge_range("A" + str(row) +":I" + str(row), "CASUAL", employee_status)
        row = row + 1
        
        for index, user_with_saln in enumerate(users_with_saln_and_casual, start=0):
            ws.write(row-1, 0, index+1, table_body_a)
            ws.merge_range("B"+str(row) + ":C"+str(row), user_with_saln.last_name, table_body)
            ws.write(row-1, 3, f"{user_with_saln.first_name}, {user_with_saln.name_extn}" if user_with_saln.name_extn != 'N/A' else user_with_saln.first_name, table_body)
            ws.write(row-1, 4, user_with_saln.middle_name, table_body)
            ws.write(row-1, 5, user_with_saln.tin, table_body_a)
            ws.merge_range("G"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
            ws.write(row-1, 8, user_with_saln.saln_summary.networth, table_body_a)
            
            row = row + 1

    ws.write("B" + str(row+1), "Total Number of Filers:", cell_format_cambria_total)
    ws.write("B" + str(row+2), "Total Number of Personnel Complement:", cell_format_cambria_total)
    ws.write("B" + str(row+5), "Prepared By:", cell_format_cambria_italic)
    
    ws.write("E" + str(row+1), len(users_with_saln), cell_format_cambria_bold_signatory)
    ws.write("E" + str(row+2), len(users_with_saln), cell_format_cambria_bold_signatory)
    
    ws.merge_range("B"+str(row+9) +":D"+str(row+9), "FRANCIS CARLO L. ZACARIAS", cell_format_cambria_bold_signatory)
    ws.merge_range("B"+str(row+10) +":D"+str(row+10), "Person-in-charge of SALN", cell_format_cambria_italic_signatory)

    ws.write("B" + str(row+12), "Position:", cell_format_cambria_contact)
    ws.write("B" + str(row+13), "Email Address:", cell_format_cambria_contact)
    ws.write("B" + str(row+14), "Contact No.:", cell_format_cambria_contact)
    ws.write("B" + str(row+15), "Date:", cell_format_cambria_contact)

    ws.write("C" + str(row+12), "Industrial Relations Management/Development Officer C", cell_format_cambria_contact)
    ws.write("C" + str(row+13), "pimo.adm02@gmail.com", cell_format_cambria_contact)
    ws.write("C" + str(row+14), "(075) 505-1550", cell_format_cambria_contact)
    ws.write("C" + str(row+15), "January 11, 2023", cell_format_cambria_contact)

    ws.write("G" + str(row+5), "Noted By:", cell_format_cambria_italic)
    ws.merge_range("G"+str(row+9) +":H"+str(row+9), "ENGR. JOHN N. MOLANO", cell_format_cambria_bold_signatory)
    ws.merge_range("G"+str(row+10) +":H"+str(row+10), "Acting Division Manager, Pangasinan IMO", cell_format_cambria_italic_signatory)

    ws.write("G" + str(row+12), "Position:", cell_format_cambria_contact)
    ws.write("G" + str(row+13), "Email Address:", cell_format_cambria_contact)
    ws.write("G" + str(row+14), "Contact No.:", cell_format_cambria_contact)
    ws.write("G" + str(row+15), "Date:", cell_format_cambria_contact)

    ws.write("H" + str(row+12), "Acting Division Manager", cell_format_cambria_contact)
    ws.write("H" + str(row+13), "pimo.oimofficial@gmail.com", cell_format_cambria_contact)
    ws.write("H" + str(row+14), "(075) 632-2775", cell_format_cambria_contact)
    ws.write("H" + str(row+15), "January 11, 2023", cell_format_cambria_contact)

    ws.set_row(row+12, 13)
    ws.set_row(row+13, 13)
    ws.set_row(row+14, 13)
    ws.set_row(row+15, 13)

    

    ws = workbook.add_worksheet("JOB ORDER")

    ws.set_paper(9)
    ws.set_landscape()
    ws.center_horizontally()
    ws.set_margins(left=0.13, right=0.13, top=1.21, bottom=1.21)

    ws.repeat_rows(0, 4)


    ws.set_header('&L&G', {'image_left': os.path.join(current_app.root_path, 'static/saln', 'landscape-header-cropped.png')}, margin=0)
    ws.set_footer('&R&G', {'image_right': os.path.join(current_app.root_path, 'static/saln', 'landscape-footer-cropped.png')}, margin=0)


    ws.merge_range("A1:I1", "SUMMARY LIST OF FILERS", cell_header_format)
    ws.merge_range("A2:I2", "Statement of Assets, Liabilities and Net Worth", cell_subheader)
    ws.merge_range("A3:I3", "Calendar Year 2023", cell_subheader)

    ws.merge_range("A4:A5", "NO.", table_header_format)
    ws.merge_range("B4:E4", "NAME OF EMPLOYEE", table_header_format)
    ws.merge_range("B5:C5", "Last Name", table_subheader_format)
    ws.write("D5", "First Name", table_subheader_format)
    ws.write("E5", "Middle Name", table_subheader_format)
    ws.merge_range("F4:F5", "TIN", table_header_format)
    ws.merge_range("G4:H5", "POSITION TITLE", table_header_format)
    # ws.merge_range("G4:G5", "NET WORTH", table_header_format)
    ws.merge_range("I4:I5", "NET WORTH", table_header_format)

    ws.merge_range("A6:I6", "JOB ORDER", employee_status)

    ws.set_row(0, 15)
    ws.set_row(1, 14.25)
    ws.set_row(2, 22.5)
    ws.set_row(3, 15, cell_format)

    ws.set_column('A:A', 4.29, cell_format_center)
    ws.set_column('B:B', 14.43, cell_format_cambria)
    ws.set_column('C:C', 0.83, cell_format_cambria)
    ws.set_column('D:D', 16.57, cell_format_cambria)
    ws.set_column('E:E', 15.29, cell_format_center)
    ws.set_column('F:F', 14.14, cell_format_cambria)
    ws.set_column('G:G', 14.86, cell_format_cambria)
    ws.set_column('H:H', 40.43, cell_format_cambria)
    ws.set_column('I:I', 19.29, cell_format_cambria)

    row = 7

    for index, user_with_saln in enumerate(jo_users_with_saln, start=0):
        ws.write(row-1, 0, index+1, table_body_a)
        ws.merge_range("B"+str(row) + ":C"+str(row), user_with_saln.last_name, table_body)
        ws.write(row-1, 3, f"{user_with_saln.first_name}, {user_with_saln.name_extn}" if user_with_saln.name_extn != 'N/A' else user_with_saln.first_name, table_body)
        ws.write(row-1, 4, user_with_saln.middle_name, table_body)
        ws.write(row-1, 5, user_with_saln.tin, table_body_a)
        ws.merge_range("G"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
        ws.write(row-1, 8, user_with_saln.saln_summary.networth, table_body_a)
        
        row = row + 1

    ws.write("B" + str(row+1), "Total Number of Filers:", cell_format_cambria_total)
    ws.write("B" + str(row+2), "Total Number of Personnel Complement:", cell_format_cambria_total)
    ws.write("B" + str(row+5), "Prepared By:", cell_format_cambria_italic)
    
    ws.write("E" + str(row+1), len(jo_users_with_saln), cell_format_cambria_bold_signatory)
    ws.write("E" + str(row+2), len(jo_users_with_saln), cell_format_cambria_bold_signatory)
    
    ws.merge_range("B"+str(row+9) +":D"+str(row+9), "FRANCIS CARLO L. ZACARIAS", cell_format_cambria_bold_signatory)
    ws.merge_range("B"+str(row+10) +":D"+str(row+10), "Person-in-charge of SALN", cell_format_cambria_italic_signatory)

    ws.write("B" + str(row+12), "Position:", cell_format_cambria_contact)
    ws.write("B" + str(row+13), "Email Address:", cell_format_cambria_contact)
    ws.write("B" + str(row+14), "Contact No.:", cell_format_cambria_contact)
    ws.write("B" + str(row+15), "Date:", cell_format_cambria_contact)

    ws.write("C" + str(row+12), "Industrial Relations Management/Development Officer C", cell_format_cambria_contact)
    ws.write("C" + str(row+13), "pimo.adm02@gmail.com", cell_format_cambria_contact)
    ws.write("C" + str(row+14), "(075) 505-1550", cell_format_cambria_contact)
    ws.write("C" + str(row+15), "January 11, 2023", cell_format_cambria_contact)

    ws.write("G" + str(row+5), "Noted By:", cell_format_cambria_italic)
    ws.merge_range("G"+str(row+9) +":H"+str(row+9), "ENGR. JOHN N. MOLANO", cell_format_cambria_bold_signatory)
    ws.merge_range("G"+str(row+10) +":H"+str(row+10), "Acting Division Manager, Pangasinan IMO", cell_format_cambria_italic_signatory)

    ws.write("G" + str(row+12), "Position:", cell_format_cambria_contact)
    ws.write("G" + str(row+13), "Email Address:", cell_format_cambria_contact)
    ws.write("G" + str(row+14), "Contact No.:", cell_format_cambria_contact)
    ws.write("G" + str(row+15), "Date:", cell_format_cambria_contact)

    ws.write("H" + str(row+12), "Acting Division Manager", cell_format_cambria_contact)
    ws.write("H" + str(row+13), "pimo.oimofficial@gmail.com", cell_format_cambria_contact)
    ws.write("H" + str(row+14), "(075) 632-2775", cell_format_cambria_contact)
    ws.write("H" + str(row+15), "January 11, 2023", cell_format_cambria_contact)

    ws.set_row(row+12, 13)
    ws.set_row(row+13, 13)
    ws.set_row(row+14, 13)
    ws.set_row(row+15, 13)
    
    workbook.close()

    buffer.seek(0)
      # Save the workbook to the specified file path
    file_path = os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers_test.xlsx')
    with open(file_path, 'wb') as file:
        file.write(buffer.getvalue())


    # wb.save(buffer)
    buffer.seek(0)  # Reset the buffer position to the beginning

    
    return send_file(
        buffer,
        download_name='summary_list_of_filers_modified.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ---------------------------------------------------------------------------- #


@salnReports.route('generate-report-on-saln', methods=['POST', 'GET'])
@login_required
def generate_report_on_saln():
    users_with_saln = User.query.join(Saln_Summary).filter(User.saln_summary != None).filter(User.status_remarks == "ACTIVE").filter(User.employment_status != "JOB ORDER").filter(User.employment_status != "CONTRACT OF SERVICE").all()
    jo_users_with_saln = User.query.join(Saln_Summary).filter(
                                                            User.saln_summary != None,
                                                            User.status_remarks == "ACTIVE",
                                                            or_(User.employment_status == "JOB ORDER", User.employment_status == "CONTRACT OF SERVICE")
                                                            ).all()

    users_with_saln_and_permanent = [user for user in users_with_saln if user.employment_status == "PERMANENT"]
    users_with_saln_and_coterm = [user for user in users_with_saln if user.employment_status == "COTERMINOUS"]
    users_with_saln_and_casual = [user for user in users_with_saln if user.employment_status == "CASUAL"]
    

    template = os.path.join(current_app.root_path, 'static/templates', 'summary_list_of_filers.xlsx')  
    new_save_path = os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers.xlsx')  

    # Create a BytesIO buffer to save the workbook
    buffer = BytesIO()

    # workbook = xlsxwriter.Workbook(os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers_test.xlsx'))
    workbook = xlsxwriter.Workbook(buffer)
    # default cell format to size 10 

    ws = workbook.add_worksheet("PERM-CAS-COT")
    ws.set_paper(9)
    # ws.set_landscape()
    ws.center_horizontally()
    ws.set_margins(left=0.13, right=0.13, top=1.3, bottom=1.21)

    ws.set_header('&L&G', {'image_left': os.path.join(current_app.root_path, 'static/saln', 'test-header.png')}, margin=0)
    # ws.set_header("&RPage &P of &N", margin=0)
    ws.set_footer('&R&G', {'image_right': os.path.join(current_app.root_path, 'static/saln', 'test-footer.png')}, margin=0)

    ws.set_print_scale(94)


    cell_format = workbook.add_format({'bold': True, 'font_name':'Cambria'})
    cell_format_cambria = workbook.add_format({'font_name':'Cambria', 'valign':'left',})
    cell_format_cambria_center = workbook.add_format({'font_name':'Cambria', 'valign':'center',})
    cell_format_cambria_total = workbook.add_format({'font_name':'Cambria', 'valign':'left', 'font_size':9})
    cell_format_cambria_contact = workbook.add_format({'font_name':'Cambria', 'valign':'left', 'font_size':9})
    cell_format_cambria_italic = workbook.add_format({'font_name':'Cambria', 'italic' : True,})
    cell_format_cambria_bold= workbook.add_format({'font_name':'Cambria', 'bold' : True,})
    cell_format_cambria_bold_border_top = workbook.add_format({'font_name':'Cambria', 'bold' : True, 'valign':'vcenter', 'align': 'center'})
    cell_format_cambria_italic_signatory = workbook.add_format({'font_name':'Cambria', 'italic' : True, 'valign': 'vcenter'})
    cell_format_cambria_bold_signatory= workbook.add_format({'font_name':'Cambria', 'bold' : True, 'valign': 'vcenter'})
    cell_format_center = workbook.add_format({'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter'})
    cell_header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'font_name':'Cambria', 'align': 'center', 'font_size': 14})
    cell_subheader = workbook.add_format({'font_name':'Cambria', 'align':'center', 'valign':'top'})
    table_header_format = workbook.add_format({'bold': True, 'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter', 'border' : 1 })
    table_subheader_format = workbook.add_format({'bold': True, 'italic' : True, 'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter', 'border' : 1})
    table_body = workbook.add_format({'valign': 'left','font_name':'Cambria', 'border' : 1, 'font_size': 9})
    table_body_a = workbook.add_format({'valign': 'center','font_name':'Cambria', 'border' : 1, 'font_size': 9})
    table_body_a_center = workbook.add_format({'valign': 'center','font_name':'Cambria', 'border' : 1, 'font_size': 9})
    table_body_a_italic = workbook.add_format({'valign': 'center','font_name':'Cambria', 'border' : 1, 'font_size': 9})

    employee_status = workbook.add_format({'bold': True,'font_name':'Cambria', 'border' : 1, 'font_size': 9})

    today = datetime.date.today()

    ws.merge_range("A1:J1", "REPORT ON\nSTATEMENT OF ASSETS, LIABILITY AND NET WORTH (SALN) FOR CY " + str(today.year - 1), cell_header_format)
   
    ws.write("A3", "Region", cell_format_cambria)
    ws.write("A4", "Name of Agency", cell_format_cambria)
    ws.write("A5", "Name of Office", cell_format_cambria)
    ws.write("A6", "Total No. of Employees", cell_format_cambria)
    ws.write("B7", "PERMANENT", cell_format_cambria)
    ws.write("B8", "COTERMINOUS", cell_format_cambria)
    ws.write("B9", "CASUAL", cell_format_cambria)
    ws.write("B10", "TOTAL", cell_format_cambria_bold)

    ws.write("D3", ":", cell_format_cambria)
    ws.write("D4", ":", cell_format_cambria)
    ws.write("D5", ":", cell_format_cambria)
    ws.write("D6", ":", cell_format_cambria)
    ws.write("D7", ":", cell_format_cambria)
    ws.write("D8", ":", cell_format_cambria)
    ws.write("D9", ":", cell_format_cambria)
    ws.write("D10", ":", cell_format_cambria_bold)


    ws.write("F3", "1", cell_format_cambria_center)
    ws.write("F4", "NATIONAL IRRIGATION ADMINISTRATION", cell_format_cambria)
    ws.write("F5", "PANGASINAN IRRIGATION MANAGEMENT OFFICE", cell_format_cambria)
    ws.write("F6", len(users_with_saln), cell_format_cambria_center)
    ws.write("F7", len(users_with_saln_and_permanent), cell_format_cambria_center)
    ws.write("F8", len(users_with_saln_and_coterm), cell_format_cambria_center)
    ws.write("F9", len(users_with_saln_and_casual), cell_format_cambria_center)
    cell_format_cambria_bold_border_top.set_top(6)
    ws.write("F10", len(users_with_saln), cell_format_cambria_bold_border_top)

    ws.write("A12", "NO.", table_header_format)
    ws.merge_range("B12:D12", "NAME OF EMPLOYEE", table_header_format)
    ws.merge_range("E12:H12", "POSITION TITLE", table_header_format)
    ws.write("I12", "JG", table_header_format)
    ws.write("J12", "REMARKS", table_header_format)

    ws.set_row(0, 41.25)
    ws.set_row(11, 45.75)

    ws.set_column('A:A', 4, cell_format_center)
    ws.set_column('B:B', 10.43, cell_format_cambria)
    ws.set_column('C:C', 14.29, cell_format_cambria)
    ws.set_column('D:D', 3, cell_format_cambria)
    ws.set_column('E:E', 3.71, cell_format_center)
    ws.set_column('F:F', 9.57, cell_format_cambria)
    ws.set_column('G:G', 16, cell_format_cambria)
    ws.set_column('H:H', 10.86, cell_format_cambria)
    ws.set_column('I:I', 4.29, cell_format_center)
    ws.set_column('J:J', 19.43, cell_format_cambria)

    row = 13

    ws.merge_range("A" + str(row) +":J" + str(row), "PERMANENT", employee_status)
    row = row + 1

    for index, user_with_saln in enumerate(users_with_saln_and_permanent, start=0):
        ws.write(row-1, 0, index+1, table_body_a)
        ws.merge_range("B"+str(row) + ":D"+str(row), user_with_saln.proper_fullname, table_body)
        ws.merge_range("E"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
        ws.write(row-1, 8, user_with_saln.job_grade, table_body_a_center)
        if user_with_saln.is_spouse_saln_filer != "unchecked":
            ws.write(row-1, 9, "Spouse is the Filer", table_body_a)
        else:
            ws.write(row-1, 9, "", table_body_a)

        row = row + 1

    if users_with_saln_and_coterm:
        ws.merge_range("A" + str(row) +":J" + str(row), "COTERMINOUS", employee_status)
        row = row + 1
        
        for index, user_with_saln in enumerate(users_with_saln_and_coterm, start=0):
            ws.write(row-1, 0, index+1, table_body_a)
            ws.merge_range("B"+str(row) + ":D"+str(row), user_with_saln.proper_fullname, table_body)
            ws.merge_range("E"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
            ws.write(row-1, 8, user_with_saln.job_grade, table_body_a_center)
            if user_with_saln.is_spouse_saln_filer != "unchecked":
                ws.write(row-1, 9, "Spouse is the Filer", table_body_a)
            else:
                ws.write(row-1, 9, "", table_body_a)

            row = row + 1


    if users_with_saln_and_casual:
        ws.merge_range("A" + str(row) +":J" + str(row), "CASUAL", employee_status)
        row = row + 1
        
        for index, user_with_saln in enumerate(users_with_saln_and_casual, start=0):
            ws.write(row-1, 0, index+1, table_body_a)
            ws.merge_range("B"+str(row) + ":D"+str(row), user_with_saln.proper_fullname, table_body)
            ws.merge_range("E"+str(row) + ":H"+str(row), user_with_saln.position_title, table_body)
            ws.write(row-1, 8, user_with_saln.job_grade, table_body_a_center)
            if user_with_saln.is_spouse_saln_filer != "unchecked":
                ws.write(row-1, 9, "Spouse is the Filer", table_body_a)
            else:
                ws.write(row-1, 9, "", table_body_a)

            row = row + 1

    cell_format_cambria_bold_signatory.set_border(0)

    ws.write("B"+str(row+1), "Prepared by:", cell_format_cambria_italic)
    ws.write("H"+str(row+1), "Certified Correct:", cell_format_cambria_italic)


    ws.write("B"+str(row+6), "FRANCIS CARLO L. ZACARIAS", cell_format_cambria_bold_signatory)
    ws.write("H"+str(row+6), "ENGR. JOHN N. MOLANO", cell_format_cambria_bold_signatory)
    ws.write("B"+str(row+7), "Industrial Relations Development/Management Officer C", cell_format_cambria_italic_signatory)
    ws.write("H"+str(row+7), "Acting Division Manager", cell_format_cambria_italic_signatory)


    
    workbook.close()

    buffer.seek(0)
      # Save the workbook to the specified file path
    file_path = os.path.join(current_app.root_path, 'static/saln', 'report_on_saln.xlsx')
    with open(file_path, 'wb') as file:
        file.write(buffer.getvalue())


    # wb.save(buffer)
    buffer.seek(0)  # Reset the buffer position to the beginning

    
    return send_file(
        buffer,
        download_name='summary_list_of_filers_modified.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ---------------------------------------------------------------------------- #


@salnReports.route('generate-officers-employees-submission', methods=['POST', 'GET'])
@login_required
def generate_officers_employees_submission():
    users_with_saln = User.query.join(Saln_Summary).filter(User.saln_summary != None).filter(User.status_remarks == "ACTIVE").filter(User.employment_status != "JOB ORDER").filter(User.employment_status != "CONTRACT OF SERVICE").all()
    jo_users_with_saln = User.query.join(Saln_Summary).filter(
                                                            User.saln_summary != None,
                                                            User.status_remarks == "ACTIVE",
                                                            or_(User.employment_status == "JOB ORDER", User.employment_status == "CONTRACT OF SERVICE")
                                                            ).all()

    users_with_saln_and_permanent = [user for user in users_with_saln if user.employment_status == "PERMANENT"]
    users_with_saln_and_coterm = [user for user in users_with_saln if user.employment_status == "COTERMINOUS"]
    users_with_saln_and_casual = [user for user in users_with_saln if user.employment_status == "CASUAL"]
    

    template = os.path.join(current_app.root_path, 'static/templates', 'summary_list_of_filers.xlsx')  
    new_save_path = os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers.xlsx')  

    # Create a BytesIO buffer to save the workbook
    buffer = BytesIO()

    # workbook = xlsxwriter.Workbook(os.path.join(current_app.root_path, 'static/saln', 'modified_summary_list_of_filers_test.xlsx'))
    workbook = xlsxwriter.Workbook(buffer)
    # default cell format to size 10 

    ws = workbook.add_worksheet("OFFICERS-AND-EMPLOYEES")
    ws.set_paper(9)
    ws.set_landscape()
    # ws.center_horizontally()
    ws.set_margins(left=0.13, right=0.13, top=1.10, bottom=1.08)

    ws.set_header('&L&G', {'image_left': os.path.join(current_app.root_path, 'static/saln', 'landscape-header-cropped.png')}, margin=0)
    # ws.set_header("&RPage &P of &N", margin=0)
    ws.set_footer('&R&G', {'image_right': os.path.join(current_app.root_path, 'static/saln', 'landscape-footer-cropped.png')}, margin=0)

    ws.set_print_scale(92)

    ws.repeat_rows(0)


    cell_format = workbook.add_format({'bold': True, 'font_name':'Cambria', 'valign':'vcenter', 'font_size': 9})
    cell_format_cambria = workbook.add_format({'font_name':'Cambria', 'valign':'left',})
    cell_format_cambria_center = workbook.add_format({'font_name':'Cambria', 'valign':'center',})
    cell_format_cambria_total = workbook.add_format({'font_name':'Cambria', 'valign':'left', 'font_size':9})
    cell_format_cambria_contact = workbook.add_format({'font_name':'Cambria', 'valign':'left', 'font_size':9})
    cell_format_cambria_italic = workbook.add_format({'font_name':'Cambria', 'italic' : True, 'valign':'vcenter'})
    cell_format_cambria_bold= workbook.add_format({'font_name':'Cambria', 'bold' : True, 'valign':'vcenter'})
    cell_format_cambria_bold_border_top = workbook.add_format({'font_name':'Cambria', 'bold' : True, 'valign':'vcenter', 'align': 'center'})
    cell_format_cambria_italic_signatory = workbook.add_format({'font_name':'Cambria', 'italic' : True, 'valign': 'vcenter'})
    cell_format_cambria_bold_signatory= workbook.add_format({'font_name':'Cambria', 'bold' : True, 'valign': 'vcenter'})
    
    cell_format_center = workbook.add_format({'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter'})
    
    cell_header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'font_name':'Cambria', 'align': 'center', 'valign':'vcenter', 'font_size': 14})
    cell_subheader = workbook.add_format({'font_name':'Cambria', 'align':'center', 'valign':'top'})
    table_header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter', 'border' : 1, 'font_size': 9 })
    table_subheader_format = workbook.add_format({'bold': True, 'italic' : True, 'font_name':'Cambria', 'align': 'center', 'valign': 'vcenter', 'border' : 1})
    table_body = workbook.add_format({'valign': 'left','font_name':'Cambria', 'border' : 1, 'font_size': 9})
    table_body_a = workbook.add_format({'valign': 'vcenter','font_name':'Cambria', 'border' : 1, 'font_size': 9, 'align':'center', 'bold':True})
    table_body_a_center = workbook.add_format({'valign': 'vcenter', 'align':'center','font_name':'Cambria', 'border' : 1, 'font_size': 9})
    table_body_a_italic = workbook.add_format({'valign': 'vcenter','font_name':'Cambria', 'border' : 1, 'font_size': 9})

    employee_status = workbook.add_format({'bold': True,'font_name':'Cambria', 'border' : 1, 'font_size': 9})

    today = datetime.date.today()

    ws.merge_range("A1:N1", "OFFICERS AND EMPLOYEES' SUBMISSION OF SALN", cell_header_format)

    ws.write("A2", "BUEREAU / ATTACHED AGENCY / DELIVERY UNIT", table_header_format)
    ws.write("B2", "Total Number of Employees Covered by RA6713", table_header_format)
    ws.write("C2", "Number of Employees Filed SALN", table_header_format)
    ws.write("D2", "PERCENTAGE OF COMPLIANCE (%)", table_header_format)
    ws.write("E2", "NO.", table_header_format)
    ws.merge_range("F2:H2", "NAME OF EMPLOYEE", table_header_format)
    ws.merge_range("I2:L2", "POSITION TITLE", table_header_format)
    ws.write("M2", "JG", table_header_format)
    ws.write("N2", "REMARKS", table_header_format)

    ws.set_row(0, 41.25)
    ws.set_row(1, 57.75)
    ws.set_row(2, 26.25)

    ws.set_column('A:A', 15.14, cell_format_center)
    ws.set_column('B:B', 12.86, cell_format_cambria)
    ws.set_column('C:C', 12.86, cell_format_cambria)
    ws.set_column('D:D', 12.86, cell_format_cambria)
    ws.set_column('E:E', 4, cell_format_center)
    ws.set_column('F:F', 10.43, cell_format_cambria)
    ws.set_column('G:G', 14.29, cell_format_cambria)
    ws.set_column('H:H', 3, cell_format_cambria)
    ws.set_column('I:I', 3.71, cell_format_center)
    ws.set_column('J:J', 9.57, cell_format_cambria)
    ws.set_column('K:K', 16, cell_format_cambria)
    ws.set_column('L:L', 10.86, cell_format_cambria)
    ws.set_column('M:M', 4.29, cell_format_cambria)
    ws.set_column('N:N', 19.43, cell_format_cambria)

    ws.write('A3', "PANGASINAN IMO", table_body_a)
    ws.write('B3', len(users_with_saln), table_body_a)
    ws.write('C3', len(users_with_saln), table_body_a)
    ws.write('D3', "100%", table_body_a)
    ws.merge_range("E3:N3", "Employees who submitted duly accomplished SALN", table_header_format)

    row = 4

    ws.merge_range("E" + str(row) +":N" + str(row), "PERMANENT", employee_status)
    row = row + 1

    for index, user_with_saln in enumerate(users_with_saln_and_permanent, start=0):
        ws.write(row-1, 4, index+1, table_body_a)
        ws.merge_range("F"+str(row) + ":H"+str(row), user_with_saln.proper_fullname, table_body)
        ws.merge_range("I"+str(row) + ":L"+str(row), user_with_saln.position_title, table_body)
        ws.write(row-1, 12, user_with_saln.job_grade, table_body_a_center)
        if user_with_saln.is_spouse_saln_filer != "unchecked":
            ws.write(row-1, 13, "Spouse is the Filer", table_body_a_center)
        else:
            ws.write(row-1, 13, "", table_body_a_center)

        row = row + 1

    if users_with_saln_and_coterm:
        ws.merge_range("E" + str(row) +":N" + str(row), "COTERMINOUS", employee_status)
        row = row + 1

        for index, user_with_saln in enumerate(users_with_saln_and_coterm, start=0):
            ws.write(row-1, 4, index+1, table_body_a)
            ws.merge_range("F"+str(row) + ":H"+str(row), user_with_saln.proper_fullname, table_body)
            ws.merge_range("I"+str(row) + ":L"+str(row), user_with_saln.position_title, table_body)
            ws.write(row-1, 12, user_with_saln.job_grade, table_body_a_center)
            if user_with_saln.is_spouse_saln_filer != "unchecked":
                ws.write(row-1, 13, "Spouse is the Filer", table_body_a_center)
            else:
                ws.write(row-1, 13, "", table_body_a_center)

            row = row + 1


    if users_with_saln_and_casual:
        ws.merge_range("E" + str(row) +":N" + str(row), "CASUAL", employee_status)
        row = row + 1

        for index, user_with_saln in enumerate(users_with_saln_and_casual, start=0):
            ws.write(row-1, 4, index+1, table_body_a)
            ws.merge_range("F"+str(row) + ":H"+str(row), user_with_saln.proper_fullname, table_body)
            ws.merge_range("I"+str(row) + ":L"+str(row), user_with_saln.position_title, table_body)
            ws.write(row-1, 12, user_with_saln.job_grade, table_body_a_center)
            if user_with_saln.is_spouse_saln_filer != "unchecked":
                ws.write(row-1, 13, "Spouse is the Filer", table_body_a_center)
            else:
                ws.write(row-1, 13, "", table_body_a_center)

            row = row + 1

    ws.merge_range("A4:A" + str(row-1),"", table_body)
    ws.merge_range("B4:B" + str(row-1),"", table_body)
    ws.merge_range("C4:C" + str(row-1),"", table_body)
    ws.merge_range("D4:D" + str(row-1),"", table_body)

    row = row + 2
    
    ws.write("A"+str(row), "BUEREAU / ATTACHED AGENCY / DELIVERY UNIT", table_header_format)
    ws.write("B"+str(row), "Total Number of Employees Covered by RA6713", table_header_format)
    ws.write("C"+str(row), "Number of Employees Filed SALN", table_header_format)
    ws.write("D"+str(row), "PERCENTAGE OF COMPLIANCE (%)", table_header_format)
    ws.write("E"+str(row), "NO.", table_header_format)
    ws.merge_range("F"+str(row)+":H"+str(row), "NAME OF EMPLOYEE", table_header_format)
    ws.merge_range("I"+str(row)+":L"+str(row), "POSITION TITLE", table_header_format)
    ws.write("M"+str(row), "JG", table_header_format)
    ws.write("N"+str(row), "REMARKS", table_header_format)

    ws.set_row(row-1, 57.75)
    ws.set_row(row, 26.25)

    row = row + 1
    ws.write('A'+str(row), "PANGASINAN IMO", table_body_a)
    ws.write('B'+str(row), "0", table_body_a)
    ws.write('C'+str(row), "0", table_body_a)
    ws.write('D'+str(row), "", table_body_a)
    ws.merge_range("E"+str(row)+":N"+str(row), "Employees who submitted duly accomplished SALN", table_header_format)

    row = row + 1

    ws.write("A"+str(row), "", table_body)
    ws.write("B"+str(row), "", table_body)
    ws.write("C"+str(row), "", table_body)
    ws.write("D"+str(row), "", table_body)
    ws.merge_range("E"+str(row)+":N"+str(row), "NONE", table_header_format)


    row = row + 2

    cell_format_cambria_bold_signatory.set_border(0)

    ws.write("B"+str(row+1), "Prepared by:", cell_format_cambria_italic)
    ws.write("J"+str(row+1), "Certified Correct:", cell_format_cambria_italic)

    ws.write("B"+str(row+6), "FRANCIS CARLO L. ZACARIAS", cell_format_cambria_bold_signatory)
    ws.write("J"+str(row+6), "ENGR. JOHN N. MOLANO", cell_format_cambria_bold_signatory)
    ws.write("B"+str(row+7), "Industrial Relations Development/Management Officer C", cell_format_cambria_italic_signatory)
    ws.write("J"+str(row+7), "Acting Division Manager", cell_format_cambria_italic_signatory)


    # ws.merge_range("E" + str(row) +":N" + str(row), "PERMANENT", employee_status)

    
    workbook.close()

    buffer.seek(0)
      # Save the workbook to the specified file path
    file_path = os.path.join(current_app.root_path, 'static/saln', 'report_on_saln.xlsx')
    with open(file_path, 'wb') as file:
        file.write(buffer.getvalue())


    # wb.save(buffer)
    buffer.seek(0)  # Reset the buffer position to the beginning

    
    return send_file(
        buffer,
        download_name='summary_list_of_filers_modified.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@salnReports.route('generate-transmittal', methods=['POST', 'GET'])
@login_required
def generate_transmittal():
    template = os.path.join(current_app.root_path, 'static/saln', 'saln_transmittal.docx')

    document = from_template(template)
    document.seek(0)
    
    return send_file(
        document, mimetype='application/vnd.openxmlformats-'
        'officedocument.wordprocessingml.document', as_attachment=True,
        attachment_filename='SALN_Transmittal.docx')



def get_context(id):
    """ You can generate your context separately since you may deal with a lot 
        of documents. You can carry out computations, etc in here and make the
        context look like the sample below.
    """

    users = db.session.query(User).all()

    return users

def from_template(template):
    target_file = BytesIO()

    template = DocxTemplate(template)
    context = get_context(emp_id)  # gets the context used to render the document
    
    target_file = BytesIO()
    template.render(context, autoescape=True)
    template.save(target_file)

    return target_file